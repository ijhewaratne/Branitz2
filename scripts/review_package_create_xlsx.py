"""
Best-effort creation of feedback_template.xlsx for the Phase 5 review package.

This avoids adding a hard dependency on openpyxl; if openpyxl is unavailable,
we simply keep the CSV feedback template.
"""

from __future__ import annotations

import sys
from pathlib import Path


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: python scripts/review_package_create_xlsx.py <out_dir>")
        return 2

    out_dir = Path(sys.argv[1])
    csv_path = out_dir / "feedback_template.csv"
    xlsx_path = out_dir / "feedback_template.xlsx"

    if not csv_path.exists():
        print(f"feedback_template.csv not found: {csv_path}")
        return 1

    try:
        import openpyxl  # type: ignore
    except Exception:
        # No dependency available; keep CSV only.
        print("openpyxl not available; leaving feedback_template.csv only.")
        return 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback"

    # Copy CSV header + one blank row.
    lines = csv_path.read_text(encoding="utf-8").splitlines()
    for r_idx, line in enumerate(lines, start=1):
        cells = line.split(",")
        for c_idx, value in enumerate(cells, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Light formatting
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    wb.save(xlsx_path)
    print(f"Wrote: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

